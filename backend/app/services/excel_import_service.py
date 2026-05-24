"""
Excel 批量导入服务
用户导入、数据资产导入（openpyxl）
"""
import uuid
import logging
from datetime import datetime, timezone
from typing import Optional, List

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.user import User, Organization, Department
from app.models.data_asset import DataAsset
from app.schemas.import_data import ImportResult, UserImportRow, AssetImportRow
from app.exceptions import DataValidationError

logger = logging.getLogger(__name__)

# 允许的 Excel MIME 类型
ALLOWED_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",
}

# 最大行数限制
MAX_ROWS = 10000


async def import_users_from_excel(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
) -> ImportResult:
    """
    从 Excel 文件批量导入用户

    Args:
        db: 数据库会话
        file_content: Excel 文件内容（bytes）
        filename: 文件名

    Returns:
        导入结果
    """
    rows = _parse_excel_file(file_content, filename, "users")
    if not rows:
        return ImportResult(total_rows=0, success_count=0, error_count=0)

    success_count = 0
    error_count = 0
    errors: List[dict] = []
    imported_ids: List[str] = []

    for row_idx, row_data in enumerate(rows, start=2):  # Excel 行号从2开始（第1行是标题）
        try:
            # 解析行数据
            import_row = UserImportRow(
                username=str(row_data.get("username", "")),
                email=row_data.get("email"),
                phone=row_data.get("phone"),
                role=str(row_data.get("role", "user")),
                organization_name=row_data.get("organization_name"),
                department_name=row_data.get("department_name"),
            )

            if not import_row.username.strip():
                errors.append({"row": row_idx, "error": "用户名不能为空"})
                error_count += 1
                continue

            # 检查用户名是否已存在
            existing = await db.execute(
                select(User).where(User.username == import_row.username)
            )
            if existing.scalar_one_or_none():
                errors.append({"row": row_idx, "error": f"用户名已存在: {import_row.username}"})
                error_count += 1
                continue

            # 查找组织
            org_id = None
            if import_row.organization_name:
                org_result = await db.execute(
                    select(Organization).where(
                        Organization.name == import_row.organization_name
                    )
                )
                org = org_result.scalar_one_or_none()
                if org:
                    org_id = org.id

            # 创建用户
            user = User(
                username=import_row.username,
                email=import_row.email or "",
                phone=import_row.phone or "",
                role=import_row.role,
                organization_id=org_id,
                status="active",
            )
            db.add(user)
            await db.flush()

            imported_ids.append(str(user.id))
            success_count += 1

        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})
            error_count += 1
            logger.warning(f"Excel import row {row_idx} failed: {e}")

    await db.commit()
    logger.info(
        f"用户批量导入完成: total={len(rows)}, success={success_count}, errors={error_count}"
    )
    return ImportResult(
        total_rows=len(rows),
        success_count=success_count,
        error_count=error_count,
        errors=errors,
        imported_ids=imported_ids,
    )


async def import_assets_from_excel(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    owner_id: Optional[str] = None,
) -> ImportResult:
    """
    从 Excel 文件批量导入数据资产

    Args:
        db: 数据库会话
        file_content: Excel 文件内容（bytes）
        filename: 文件名
        owner_id: 资产拥有者 ID

    Returns:
        导入结果
    """
    rows = _parse_excel_file(file_content, filename, "assets")
    if not rows:
        return ImportResult(total_rows=0, success_count=0, error_count=0)

    success_count = 0
    error_count = 0
    errors: List[dict] = []
    imported_ids: List[str] = []

    for row_idx, row_data in enumerate(rows, start=2):
        try:
            import_row = AssetImportRow(
                name=str(row_data.get("name", "")),
                description=row_data.get("description"),
                source_type=str(row_data.get("source_type", "file")),
                format=row_data.get("format"),
                classification=row_data.get("classification"),
                security_level=str(row_data.get("security_level", "public")),
                tags=row_data.get("tags"),
            )

            if not import_row.name.strip():
                errors.append({"row": row_idx, "error": "资产名称不能为空"})
                error_count += 1
                continue

            asset = DataAsset(
                name=import_row.name,
                description=import_row.description or "",
                source_type=import_row.source_type,
                format=import_row.format or "",
                classification=import_row.classification or "",
                security_level=import_row.security_level,
                owner_id=uuid.UUID(owner_id) if owner_id else None,
                status="active",
            )
            db.add(asset)
            await db.flush()

            imported_ids.append(str(asset.id))
            success_count += 1

        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})
            error_count += 1
            logger.warning(f"Asset import row {row_idx} failed: {e}")

    await db.commit()
    logger.info(
        f"资产批量导入完成: total={len(rows)}, success={success_count}, errors={error_count}"
    )
    return ImportResult(
        total_rows=len(rows),
        success_count=success_count,
        error_count=error_count,
        errors=errors,
        imported_ids=imported_ids,
    )


def _parse_excel_file(
    file_content: bytes,
    filename: str,
    import_type: str,
) -> List[dict]:
    """
    解析 Excel 文件为字典列表

    Args:
        file_content: 文件内容
        filename: 文件名
        import_type: 导入类型 (users/assets)

    Returns:
        行数据字典列表
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is not installed. Run: pip install openpyxl")
        raise DataValidationError(message="Excel 解析库未安装，请联系管理员")

    if not filename.endswith((".xlsx", ".xls")):
        raise DataValidationError(message="仅支持 .xlsx / .xls 格式文件")

    try:
        workbook = openpyxl.load_workbook(
            filename=__import__("io").BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
    except Exception as e:
        raise DataValidationError(message=f"Excel 文件解析失败: {e}")

    sheet = workbook.active
    if sheet is None:
        raise DataValidationError(message="Excel 文件无有效工作表")

    # 读取表头
    headers = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        value = cell.value
        if value is None:
            break
        headers.append(str(value).strip().lower())

    if not headers:
        raise DataValidationError(message="Excel 文件表头为空")

    # 验证必要列
    required_columns = _get_required_columns(import_type)
    missing = [col for col in required_columns if col not in headers]
    if missing:
        raise DataValidationError(
            message=f"缺少必要列: {', '.join(missing)}",
            data={"required": required_columns, "found": headers},
        )

    # 读取数据行
    rows: List[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue  # 跳过空行

        row_dict = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                value = row[idx]
                row_dict[header] = str(value) if value is not None else ""
            else:
                row_dict[header] = ""
        rows.append(row_dict)

        if len(rows) > MAX_ROWS:
            raise DataValidationError(
                message=f"超过最大行数限制 ({MAX_ROWS})，请分批导入"
            )

    workbook.close()
    logger.info(f"Excel parsed: {filename}, rows={len(rows)}, headers={headers}")
    return rows


def _get_required_columns(import_type: str) -> List[str]:
    """获取导入类型对应的必要列"""
    if import_type == "users":
        return ["username"]
    elif import_type == "assets":
        return ["name"]
    return []
