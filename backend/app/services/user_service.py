"""
用户管理服务
用户CRUD + 批量Excel导入 + 角色管理 + 密码重置 + 状态管理(active/inactive/locked)
"""
import uuid
import logging
import io
from datetime import datetime, timezone
from typing import Optional

from sqlalchemy import select, and_, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.user import User, Organization
from app.schemas.user import UserCreate, UserUpdate, UserResponse
from app.schemas.common import PaginatedResponse
from app.utils.pagination import PaginationParams, paginate_query
from app.exceptions import (
    UserNotFoundError, DataNotFoundError, DataAlreadyExistsError,
    DataValidationError, OpsError, PermissionDeniedError,
)
# 使用与 auth_service 一致的 SM3 密码哈希
from app.core.security import hash_password as _hash_password

logger = logging.getLogger(__name__)

# 支持的操作角色列表（第二层：用户级）
VALID_ROLES = {
    "system_admin", "org_admin", "data_steward", "data_subscriber",
    "product_developer", "product_publisher", "demand_manager",
    "approver", "operator", "auditor", "security_admin", "user",
    # 兼容旧角色
    "admin", "data_admin", "compute_user",
}

# 支持的生态角色列表（第一层：组织级）
VALID_ECO_ROLES = {
    "data_provider", "data_consumer", "data_intermediary",
    "data_trustee", "data_developer", "space_operator", "regulator", "hybrid",
}

# 支持的用户状态
VALID_STATUSES = {"active", "inactive", "locked"}

# 登录失败上限（超过则锁定）
MAX_LOGIN_FAIL_COUNT = 5

# 默认密码（批量导入时使用）
DEFAULT_TEMPORARY_PASSWORD = "Energy@2024!"


async def list_users(
    db: AsyncSession,
    params: PaginationParams,
    role: Optional[str] = None,
    status: Optional[str] = None,
    organization_id: Optional[str] = None,
    keyword: Optional[str] = None,
) -> PaginatedResponse:
    """
    用户列表（分页 + 过滤）

    Args:
        db: 数据库会话
        params: 分页参数
        role: 角色过滤
        status: 状态过滤
        organization_id: 组织过滤
        keyword: 关键词搜索（username/email/phone）

    Returns:
        分页用户列表
    """
    query = select(User)

    # 条件过滤
    if role:
        query = query.where(User.role == role)
    if status:
        query = query.where(User.status == status)
    if organization_id:
        query = query.where(User.organization_id == uuid.UUID(organization_id))
    if keyword:
        keyword_pattern = f"%{keyword}%"
        query = query.where(
            (User.username.ilike(keyword_pattern))
            | (User.email.ilike(keyword_pattern))
            | (User.phone.ilike(keyword_pattern))
        )

    result = await paginate_query(db, query, params, UserResponse)
    return result


async def create_user(
    db: AsyncSession,
    request: UserCreate,
    created_by: str = "",
) -> UserResponse:
    """
    创建用户

    Args:
        db: 数据库会话
        request: 创建请求
        created_by: 创建人 ID

    Returns:
        创建后的用户信息
    """
    # 验证角色
    if request.role not in VALID_ROLES:
        raise DataValidationError(
            message=f"无效角色: {request.role}，支持的角色: {', '.join(VALID_ROLES)}",
            data={"valid_roles": list(VALID_ROLES)},
        )

    # 检查用户名唯一性
    existing = await db.execute(
        select(User).where(User.username == request.username)
    )
    if existing.scalar_one_or_none():
        raise DataAlreadyExistsError(message=f"用户名已存在: {request.username}")

    # 检查邮箱唯一性
    if request.email:
        existing_email = await db.execute(
            select(User).where(User.email == request.email)
        )
        if existing_email.scalar_one_or_none():
            raise DataAlreadyExistsError(message=f"邮箱已被使用: {request.email}")

    # 检查组织是否存在
    org_result = await db.execute(
        select(Organization).where(Organization.id == uuid.UUID(request.organization_id))
    )
    org = org_result.scalar_one_or_none()
    if not org:
        raise DataNotFoundError(message=f"组织不存在: {request.organization_id}")

    # 创建用户
    user = User(
        username=request.username,
        password_hash=_hash_password(request.password),
        email=request.email,
        phone=request.phone,
        role=request.role,
        organization_id=uuid.UUID(request.organization_id),
        department_id=uuid.UUID(request.department_id) if request.department_id else None,
        status="active",
        mfa_enabled=False,
        login_fail_count=0,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)

    logger.info(f"用户创建成功: {user.username} (ID: {user.id}), 创建人: {created_by}")
    return UserResponse.model_validate(user)


async def get_user(
    db: AsyncSession,
    user_id: str,
) -> UserResponse:
    """
    获取用户详情

    Args:
        db: 数据库会话
        user_id: 用户 ID

    Returns:
        用户信息
    """
    result = await db.execute(
        select(User).where(User.id == uuid.UUID(user_id))
    )
    user = result.scalar_one_or_none()
    if not user:
        raise UserNotFoundError(message=f"用户不存在: {user_id}")
    return UserResponse.model_validate(user)


async def update_user(
    db: AsyncSession,
    user_id: str,
    request: UserUpdate,
) -> UserResponse:
    """
    更新用户信息

    Args:
        db: 数据库会话
        user_id: 用户 ID
        request: 更新请求

    Returns:
        更新后的用户信息
    """
    result = await db.execute(
        select(User).where(User.id == uuid.UUID(user_id))
    )
    user = result.scalar_one_or_none()
    if not user:
        raise UserNotFoundError(message=f"用户不存在: {user_id}")

    # 验证角色
    if request.role is not None and request.role not in VALID_ROLES:
        raise DataValidationError(
            message=f"无效角色: {request.role}",
            data={"valid_roles": list(VALID_ROLES)},
        )

    # 验证状态
    if request.status is not None and request.status not in VALID_STATUSES:
        raise DataValidationError(
            message=f"无效状态: {request.status}",
            data={"valid_statuses": list(VALID_STATUSES)},
        )

    # 检查邮箱唯一性
    if request.email and request.email != user.email:
        existing = await db.execute(
            select(User).where(User.email == request.email)
        )
        if existing.scalar_one_or_none():
            raise DataAlreadyExistsError(message=f"邮箱已被使用: {request.email}")

    # 更新字段
    update_data = request.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field == "department_id" and value:
            value = uuid.UUID(value)
        setattr(user, field, value)

    await db.commit()
    await db.refresh(user)

    logger.info(f"用户更新成功: {user.username} (ID: {user.id})")
    return UserResponse.model_validate(user)


async def delete_user(
    db: AsyncSession,
    user_id: str,
) -> dict:
    """
    删除用户（软删除，设置状态为 inactive）

    Args:
        db: 数据库会话
        user_id: 用户 ID

    Returns:
        删除结果
    """
    result = await db.execute(
        select(User).where(User.id == uuid.UUID(user_id))
    )
    user = result.scalar_one_or_none()
    if not user:
        raise UserNotFoundError(message=f"用户不存在: {user_id}")

    # 软删除：设置为 inactive
    user.status = "inactive"
    await db.commit()

    logger.info(f"用户已删除（软删除）: {user.username} (ID: {user.id})")
    return {"user_id": str(user.id), "status": "deleted", "deleted_at": datetime.now(timezone.utc).isoformat()}


async def batch_import_users(
    db: AsyncSession,
    file_content: bytes,
    filename: str = "users.xlsx",
    created_by: str = "",
) -> dict:
    """
    批量 Excel 导入用户

    解析 Excel 文件中的用户列表并批量创建。
    Excel 列: username, password, email, phone, role, organization_id, department_id

    Args:
        db: 数据库会话
        file_content: Excel 文件字节内容
        filename: 文件名
        created_by: 操作人 ID

    Returns:
        导入结果统计
    """
    # 解析 Excel（使用 openpyxl）
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active
    except Exception as e:
        raise DataValidationError(
            message=f"Excel 文件解析失败: {str(e)}",
        )

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    success_count = 0
    fail_count = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        try:
            if not row or not row[0]:
                continue

            username = str(row[0]).strip()
            password = str(row[1]).strip() if row[1] else DEFAULT_TEMPORARY_PASSWORD
            email = str(row[2]).strip() if row[2] else None
            phone = str(row[3]).strip() if row[3] else None
            role = str(row[4]).strip() if row[4] else "user"
            organization_id = str(row[5]).strip() if row[5] else None
            department_id = str(row[6]).strip() if row[6] else None

            if not organization_id:
                raise DataValidationError(message="组织 ID 不能为空")

            if role not in VALID_ROLES:
                raise DataValidationError(message=f"无效角色: {role}")

            # 检查用户名唯一性
            existing = await db.execute(
                select(User).where(User.username == username)
            )
            if existing.scalar_one_or_none():
                raise DataAlreadyExistsError(message=f"用户名已存在: {username}")

            user = User(
                username=username,
                password_hash=_hash_password(password),
                email=email,
                phone=phone,
                role=role,
                organization_id=uuid.UUID(organization_id),
                department_id=uuid.UUID(department_id) if department_id else None,
                status="active",
                mfa_enabled=False,
                login_fail_count=0,
            )
            db.add(user)
            success_count += 1

        except Exception as e:
            fail_count += 1
            errors.append({"row": idx, "error": str(e)})
            logger.warning(f"批量导入第 {idx} 行失败: {e}")

    await db.commit()

    logger.info(
        f"批量导入完成: 成功 {success_count}, 失败 {fail_count}, 操作人: {created_by}"
    )

    return {
        "total_rows": len(rows),
        "success_count": success_count,
        "fail_count": fail_count,
        "errors": errors[:50],
        "imported_at": datetime.now(timezone.utc).isoformat(),
    }


async def reset_password(
    db: AsyncSession,
    user_id: str,
    new_password: Optional[str] = None,
    reset_by: str = "",
) -> dict:
    """
    重置用户密码

    Args:
        db: 数据库会话
        user_id: 用户 ID
        new_password: 新密码（为空则使用默认临时密码）
        reset_by: 操作人 ID

    Returns:
        重置结果
    """
    result = await db.execute(
        select(User).where(User.id == uuid.UUID(user_id))
    )
    user = result.scalar_one_or_none()
    if not user:
        raise UserNotFoundError(message=f"用户不存在: {user_id}")

    # 生成新密码
    password = new_password or DEFAULT_TEMPORARY_PASSWORD
    user.password_hash = _hash_password(password)

    # 重置登录失败计数并解锁
    user.login_fail_count = 0
    user.locked_until = None
    if user.status == "locked":
        user.status = "active"

    await db.commit()

    logger.info(f"用户密码已重置: {user.username} (ID: {user.id}), 操作人: {reset_by}")
    return {
        "user_id": str(user.id),
        "reset_at": datetime.now(timezone.utc).isoformat(),
        "is_temporary": new_password is None,
    }
